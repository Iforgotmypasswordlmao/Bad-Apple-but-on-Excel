from PIL import Image
import cv2
from openpyxl import Workbook
from openpyxl.styles import PatternFill


VIDEO_FILENAME = "Bad Apple Video.mp4"
BadAppleVideo = cv2.VideoCapture(VIDEO_FILENAME)
EXCEL_FILENAME = "Bad Apple Excel.xlsx"
BadAppleWorkBook = Workbook()

Factor = 6
SquareSize = Factor*Factor

def main():
    success, frame = BadAppleVideo.read()
    length = int(len(frame)/Factor)
    width = int(len(frame[0])/Factor)
    success = True
    count = 0
    Framecount = 0
    print(length, width)
    while success: 
        BadAppleVideo.set(cv2.CAP_PROP_POS_MSEC,(count*1000))
        Sheet = BadAppleWorkBook.create_sheet(f"Frame-{Framecount}", Framecount)
        for L in range(length):
            for W in range(width):
                Colours = []
                for i in range(Factor):
                    for j in range(Factor):
                        x = Factor*L + i
                        y = Factor*W + j
                        Colours.append(sum(frame[x][y])/3)

                Colour = int(sum(Colours)/SquareSize)
                Colour = Colour if Colour != 84 else 255
                Colour = hex(Colour)[2:]

                
                HexColour = Colour * int(6/len(Colour))
                ExcelCell = Sheet.cell(row=L+1, column=W+1)
                ExcelCell.fill = PatternFill(start_color=HexColour, end_color=HexColour, fill_type="solid")

        print(f"Processed Frame {Framecount}")
        success, frame = BadAppleVideo.read()
        Framecount += 1
        count += 0.25
    
    BadAppleWorkBook.save(EXCEL_FILENAME)
    BadAppleVideo.release()

if __name__ == '__main__':
    main()
